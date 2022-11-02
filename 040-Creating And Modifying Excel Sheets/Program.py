#!/usr/bin/python3
# Creating and modifying excel sheets

"""
>>>> Some of the methods in the openpyxl is not available in the latest version.
        I have updated the code to include those changes.
>>>>
>>>>
>>>>
>>>>
"""
import openpyxl

# Creating a new excel workbook.
# The newly created workbook will contain a sheet called "Sheet".
wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb["Sheet"]

# Adding value into workbook
sheet["A1"] = "Apple"
sheet["A2"] = "Orange"
sheet["A3"] = "Grapes"

# Creating a new sheet.
sheet2 = wb.create_sheet()

# Changing the sheet name.
sheet2.title = "Youtube"
print(wb.sheetnames)

# Creating sheet at any position.
sheet3 = wb.create_sheet(index=0, title="My first sheet")
print(wb.sheetnames)

# Saving the excel file
wb.save("written.xlsx")

print(" ")
